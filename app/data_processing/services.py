import pickle
from collections import namedtuple
from http import HTTPStatus
from typing import Generic, TypeVar

from fastapi import HTTPException
from openpyxl.reader.excel import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.celery import celery
from app.common.services import AbstractBackgroundService
from app.config import Config
from app.data_processing.schemas import TaskCreated
from app.db import async_session_maker
from app.models import Dish, Menu, Submenu
from app.utils import is_valid_uuid

TASK_NOT_FOUND_MESSAGE = 'Task not found or still being processing.'


TaskResultType = TypeVar('TaskResultType')


class AllMenusService(AbstractBackgroundService, Generic[TaskResultType]):

    async def get_task_result(self, task_id: str) -> TaskResultType:
        task = celery.AsyncResult(task_id)
        if task.ready():
            return pickle.loads(task.result)
        raise HTTPException(
            detail=TASK_NOT_FOUND_MESSAGE,
            status_code=HTTPStatus.ACCEPTED,
        )

    async def create_task(self) -> TaskCreated:
        task = self.task_function.delay()
        return TaskCreated(task_id=task.id)


class AdminFileObjectCreationService:
    Action = namedtuple('Action', ('is_menu', 'is_submenu', 'is_dish'))

    def __init__(self):
        self._last_menu: Menu = ...
        self._last_submenu: Submenu = ...

        self._menu_action = self.Action(True, False, False)
        self._submenu_action = self.Action(False, True, False)
        self._dish_action = self.Action(False, False, True)

    async def create_missing_objects(self) -> None:
        workbook = load_workbook(Config.ADMIN_FILE_PATH, data_only=True)
        worksheet = workbook.active

        actions = {
            self._menu_action: self._handle_menu,
            self._submenu_action: self._handle_submenu,
            self._dish_action: self._handle_dish,
        }

        async with async_session_maker() as session:
            for row in worksheet.iter_rows(values_only=True):
                action = self._get_action(row)

                if self._is_action_valid(action):

                    await actions[action](row, session)

            await session.commit()

    def _get_action(self, row: tuple) -> Action:
        is_menu = is_valid_uuid(row[0])
        is_submenu = is_valid_uuid(row[1])
        is_dish = is_valid_uuid(row[2])
        return self.Action(is_menu, is_submenu, is_dish)

    def _is_action_valid(self, action):
        return action in (self._menu_action, self._submenu_action, self._dish_action)

    async def _handle_menu(self, row: tuple, session: AsyncSession) -> None:
        menu_id, title, description = row[0], row[1], row[2]
        menu = Menu(id=menu_id, title=title, description=description)
        session.add(menu)
        self._last_menu = menu

    async def _handle_submenu(self, row: tuple, session: AsyncSession) -> None:
        submenu_id, title, description = row[1], row[2], row[3]
        submenu = Submenu(id=submenu_id, title=title, description=description)
        self._last_menu.submenus.append(submenu)
        session.add(submenu)
        self._last_submenu = submenu

    async def _handle_dish(self, row: tuple, session: AsyncSession) -> None:
        dish_id, title, description, price = row[2], row[3], row[4], row[5]
        dish = Dish(id=dish_id, title=title, description=description, price=price)
        self._last_submenu.dishes.append(dish)
        session.add(dish)
