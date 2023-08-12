from abc import ABC, abstractmethod
from typing import Generic, TypeVar

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.common.repository import AbstractRepository
from app.config import Config
from app.db import async_session_maker
from app.models import Dish, Menu, Submenu
from app.utils import is_valid_uuid


class AdminFileWorksheetMixin:
    @property
    def worksheet(self):
        workbook = load_workbook(Config.ADMIN_FILE_PATH, data_only=True)
        return workbook.active


AdminFileRepositoryType = TypeVar('AdminFileRepositoryType', bound=AbstractRepository)


class AbstractAdminFileUpdatingService(AdminFileWorksheetMixin, ABC, Generic[AdminFileRepositoryType]):
    def __init__(self, repository: AdminFileRepositoryType):
        self.repository = repository

    @abstractmethod
    def create_missing_objects(self, session):
        ...


class AbstractAdminFileDeletingService(AdminFileWorksheetMixin, ABC, Generic[AdminFileRepositoryType]):
    def __init__(self, repository: AdminFileRepositoryType):
        self.repository = repository

    @abstractmethod
    def delete_redundant_objects(self, session):
        ...


class AdminFileService:

    def __init__(
            self,
            updating_service: list[AbstractAdminFileUpdatingService],
            # deletion_service: list[AbstractAdminFileDeletingService],
    ):
        self.updating_services = updating_service
        # self.deletion_services = deletion_service

    async def handle(self):
        async with async_session_maker() as session:
            for updating_service in self.updating_services:
                await updating_service.create_missing_objects(session)
            # for deletion_service in self.deletion_services:
            #     deletion_service.delete_redundant_objects()


class AdminFileMenuUpdatingService(AbstractAdminFileUpdatingService[AdminFileRepositoryType]):
    id_index = 0
    title_index = 1
    description_index = 2

    async def create_missing_objects(self, session):
        for row in self.worksheet.iter_rows(values_only=True):
            if self.is_menu_row(row):
                await self._handle(row, session)
        await session.commit()

    @classmethod
    def is_menu_row(cls, row):
        return is_valid_uuid(row[cls.id_index])

    async def _handle(self, row, session: AsyncSession):
        menu = await self.repository.get_by_id(row[self.id_index], session, orm_object=True)  # type: ignore
        if menu is None:
            await self._create_menu(row, session)
        else:
            await self._update_menu(row, menu, session)

    async def _create_menu(self, row, session):
        menu = Menu(
            id=row[self.id_index],
            title=row[self.title_index],
            description=row[self.description_index],
        )
        await self.repository.create(menu, session, commit=False)

    async def _update_menu(self, row, menu, session):
        updated_menu = Menu(
            title=row[self.title_index],
            description=row[self.description_index]
        )
        await self.repository.update(menu, updated_menu, session, commit=False)


class AdminFileSubmenuUpdatingService(AbstractAdminFileUpdatingService[AdminFileRepositoryType]):
    id_index = 1
    title_index = 2
    description_index = 3

    last_menu_id: str = ''

    async def create_missing_objects(self, session):
        for row in self.worksheet.iter_rows(values_only=True):
            if AdminFileMenuUpdatingService.is_menu_row(row):
                self.last_menu_id = row[AdminFileMenuUpdatingService.id_index]
            elif self.is_submenu_row(row):
                await self._handle(row, session)
        await session.commit()

    @classmethod
    def is_submenu_row(cls, row):
        return is_valid_uuid(row[cls.id_index])

    async def _handle(self, row, session):
        submenu = await self.repository.get_by_id(self.last_menu_id, row[self.id_index], session, orm_object=True)
        if submenu is None:
            await self._create_submenu(row, session)
        else:
            await self._update_submenu(row, submenu, session)

    async def _create_submenu(self, row, session):
        submenu = Submenu(
            id=row[self.id_index],
            title=row[self.title_index],
            description=row[self.description_index],
            menu_id=self.last_menu_id
        )
        await self.repository.create(self.last_menu_id, submenu, session, commit=False)

    async def _update_submenu(self, row, submenu, session):
        updated_submenu = Submenu(
            title=row[self.title_index],
            description=row[self.description_index],
        )
        await self.repository.update(submenu, updated_submenu, session, commit=False)


class AdminFileDishUpdatingService(AbstractAdminFileUpdatingService[AdminFileRepositoryType]):
    id_index = 2
    title_index = 3
    description_index = 4
    price_index = 5

    last_menu_id: str = ''
    last_submenu_id: str = ''

    async def create_missing_objects(self, session):
        for row in self.worksheet.iter_rows(values_only=True):
            if AdminFileMenuUpdatingService.is_menu_row(row):
                self.last_menu_id = row[AdminFileMenuUpdatingService.id_index]
            elif AdminFileSubmenuUpdatingService.is_submenu_row(row):
                self.last_submenu_id = row[AdminFileSubmenuUpdatingService.id_index]
            elif self.is_dish_row(row):
                await self._handle(row, session)
        await session.commit()

    @classmethod
    def is_dish_row(cls, row):
        return is_valid_uuid(row[cls.id_index])

    async def _handle(self, row, session):
        dish = await self.repository.get_by_id(
            self.last_menu_id,
            self.last_submenu_id,
            row[self.id_index],
            session, orm_object=True,
        )
        if dish is None:
            await self._create_dish(row, session)
        else:
            await self._update_dish(row, dish, session)

    async def _create_dish(self, row, session):
        dish = Dish(
            id=row[self.id_index],
            title=row[self.title_index],
            description=row[self.description_index],
            price=row[self.price_index],
            submenu_id=self.last_submenu_id,
        )
        await self.repository.create(self.last_submenu_id, dish, session, commit=False)

    async def _update_dish(self, row, dish, session):
        updated_dish = Dish(
            title=row[self.title_index],
            description=row[self.description_index],
        )
        await self.repository.update(dish, updated_dish, session, commit=False)
