from abc import ABC, abstractmethod
from typing import Generic
from uuid import UUID

from openpyxl.reader.excel import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import Config
from app.data_processing.services.admin import AdminServicesRepositoryType
from app.dish.cache import clear_dish_cache, clear_dish_list_cache
from app.menu.cache import clear_menu_cache, clear_menu_list_cache
from app.models import Dish, Menu, Submenu
from app.submenu.cache import clear_submenu_cache, clear_submenu_list_cache
from app.utils import is_valid_uuid


class BaseAdminUpdateService(ABC, Generic[AdminServicesRepositoryType]):
    """
    The base service that is used to create an admin
    file update service for a particular database model.
    """
    id_index: int
    visited_ids: set[str] = set()

    def __init__(self, repository: AdminServicesRepositoryType):
        self.repository = repository

    async def update(self, session: AsyncSession) -> None:
        """Synchronizes the database with the admin file."""
        workbook = load_workbook(Config.ADMIN_FILE_PATH, data_only=True)
        worksheet = workbook.active
        for row in worksheet.iter_rows(values_only=True):
            await self._create_missing_objects(row, session)
        await self._delete_irrelevant_objects(session)
        await session.commit()

    @classmethod
    def is_model_row(cls, row) -> bool:
        """
        Checks if the current row is a model row by checking
        if there is a UUID in row[index_id].
        """
        return is_valid_uuid(row[cls.id_index])

    @abstractmethod
    def _create_missing_objects(self, row: tuple, session: AsyncSession):
        """
        Creates objects that are in the admin file, but which are
        not in the database.
        """
        ...

    @abstractmethod
    def _delete_irrelevant_objects(self, session: AsyncSession):
        """
        Deletes objects which are not in the
        admin file, but which are in the database.
        """
        ...


class AdminMenuUpdateService(BaseAdminUpdateService, Generic[AdminServicesRepositoryType]):
    id_index = 0
    title_index = 1
    description_index = 2

    async def _create_missing_objects(self, row: tuple, session: AsyncSession):
        if self.is_model_row(row):
            menu_id = row[self.id_index]
            self.visited_ids.add(menu_id)
            menu = await self.repository.get_by_id(menu_id, session, orm_object=True)
            if menu is None:
                await self._create_menu(row, session)
            else:
                await self._update_menu(row, menu, session)

    async def _create_menu(self, row: tuple, session: AsyncSession):
        menu = Menu(
            id=row[self.id_index],
            title=row[self.title_index],
            description=row[self.description_index],
        )
        await self.repository.create(menu, session, commit=False)
        await clear_menu_list_cache()

    async def _update_menu(self, row: tuple, menu, session: AsyncSession):
        updated_menu = Menu(
            title=row[self.title_index],
            description=row[self.description_index]
        )
        await self.repository.update(menu, updated_menu, session, commit=False)
        await clear_menu_cache(menu.id)

    async def _delete_irrelevant_objects(self, session: AsyncSession):
        menus = await self.repository.all(session, scalar=True)
        for menu in menus:
            if str(menu.id) not in self.visited_ids:
                await self.repository.delete(menu, session, commit=False)
        self.visited_ids = set()


class AdminSubmenuUpdateService(BaseAdminUpdateService, Generic[AdminServicesRepositoryType]):
    id_index = 1
    title_index = 2
    description_index = 3

    last_menu_id: str = ''

    async def _create_missing_objects(self, row: tuple, session: AsyncSession):
        if AdminMenuUpdateService.is_model_row(row):
            self.last_menu_id = row[AdminMenuUpdateService.id_index]
        elif self.is_model_row(row):
            submenu_id = row[self.id_index]
            self.visited_ids.add(submenu_id)
            submenu = await self.repository.get_by_id(self.last_menu_id, submenu_id, session, orm_object=True)
            if submenu is None:
                await self._create_submenu(row, session)
            else:
                await self._update_submenu(row, submenu, session)

    async def _create_submenu(self, row: tuple, session: AsyncSession):
        submenu = Submenu(
            id=row[self.id_index],
            title=row[self.title_index],
            description=row[self.description_index],
            menu_id=self.last_menu_id
        )
        await self.repository.create(self.last_menu_id, submenu, session, commit=False)
        await clear_submenu_list_cache(UUID(self.last_menu_id))
        await clear_menu_cache(UUID(self.last_menu_id))

    async def _update_submenu(self, row: tuple, submenu, session: AsyncSession):
        updated_submenu = Submenu(
            title=row[self.title_index],
            description=row[self.description_index],
        )
        await self.repository.update(submenu, updated_submenu, session, commit=False)
        await clear_submenu_cache(UUID(self.last_menu_id), row[self.id_index])

    async def _delete_irrelevant_objects(self, session: AsyncSession):
        submenus = await self.repository.all(self.last_menu_id, session, scalar=True)
        for submenu in submenus:
            if str(submenu.id) not in self.visited_ids:
                await self.repository.delete(submenu, session, commit=False)
        self.visited_ids = set()


class AdminDishUpdateService(BaseAdminUpdateService, Generic[AdminServicesRepositoryType]):
    id_index = 2
    title_index = 3
    description_index = 4
    price_index = 5

    last_menu_id: str = ''
    last_submenu_id: str = ''

    async def _create_missing_objects(self, row: tuple, session: AsyncSession):
        if AdminMenuUpdateService.is_model_row(row):
            self.last_menu_id = row[AdminMenuUpdateService.id_index]
        elif AdminSubmenuUpdateService.is_model_row(row):
            self.last_submenu_id = row[AdminSubmenuUpdateService.id_index]
        elif self.is_model_row(row):
            dish_id = row[self.id_index]
            self.visited_ids.add(dish_id)
            dish = await self.repository.get_by_id(
                self.last_menu_id,
                self.last_submenu_id,
                dish_id,
                session,
                orm_object=True,
            )
            if dish is None:
                await self._create_dish(row, session)
            else:
                await self._update_dish(row, dish, session)

    async def _create_dish(self, row: tuple, session: AsyncSession):
        dish = Dish(
            id=row[self.id_index],
            title=row[self.title_index],
            description=row[self.description_index],
            price=row[self.price_index],
            submenu_id=self.last_submenu_id,
        )
        await self.repository.create(self.last_submenu_id, dish, session, commit=False)
        await clear_dish_list_cache(UUID(self.last_menu_id), row[self.id_index])
        await clear_submenu_list_cache(UUID(self.last_menu_id))
        await clear_menu_list_cache()

    async def _update_dish(self, row: tuple, dish, session: AsyncSession):
        updated_dish = Dish(
            title=row[self.title_index],
            description=row[self.description_index],
        )
        await self.repository.update(dish, updated_dish, session, commit=False)
        await clear_dish_cache(UUID(self.last_menu_id), UUID(self.last_submenu_id), row[self.id_index])

    async def _delete_irrelevant_objects(self, session: AsyncSession):
        dishes = await self.repository.all(
            self.last_menu_id,
            self.last_submenu_id,
            session,
            scalar=True,
        )
        for dish in dishes:
            if str(dish.id) not in self.visited_ids:
                await self.repository.delete(dish, session, commit=False)
        self.visited_ids = set()
